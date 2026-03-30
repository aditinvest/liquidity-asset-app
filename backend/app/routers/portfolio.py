from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, and_
from typing import Optional, List
from datetime import date, timedelta, datetime
from decimal import Decimal
import pandas as pd
from io import BytesIO
import re

from ..database import get_db
from ..models.models import PortfolioSnapshot, Asset
from ..schemas.schemas import (
    PortfolioSnapshotResponse,
    AssetResponse,
    MaturityProfileSummary,
    MaturityProfileItem,
    MaturityCalendarSummary,
    MaturityCalendarItem
)

router = APIRouter()


def parse_portfolio_excel(file: UploadFile, db: Session) -> PortfolioSnapshot:
    """Parse portfolio Excel file and store in database."""

    # Read Excel file
    contents = file.file.read()
    excel_data = pd.ExcelFile(BytesIO(contents))

    # Determine sheet names (could be "Deposito" and "Bond" or "Depo" and "Obligasi")
    deposito_sheet = "Deposito" if "Deposito" in excel_data.sheet_names else "Depo"
    bond_sheet = "Bond" if "Bond" in excel_data.sheet_names else "Obligasi"

    # Read sheets
    deposito_df = pd.read_excel(excel_data, sheet_name=deposito_sheet)
    bond_df = pd.read_excel(excel_data, sheet_name=bond_sheet)

    # Create portfolio snapshot
    snapshot = PortfolioSnapshot(
        data_source_date=date.today(),
        total_assets=Decimal('0.00')
    )
    db.add(snapshot)
    db.flush()

    total_assets = Decimal('0.00')

    # Process Deposito
    if not deposito_df.empty:
        for _, row in deposito_df.iterrows():
            # Determine OPEX/CAPEX - FundID can be "OPEX", "CAPEX", or contain "JPOIP"
            fund_id = str(row.get('FundID', ''))
            sub_fund = 'OPEX' if (fund_id.upper() == 'OPEX' or 'JPOIP' in fund_id.upper()) else 'CAPEX'

            # Determine Syariah/Konvensional
            syariah_val = str(row.get('Syariah', '')).lower()
            management_type = 'SYARIAH' if syariah_val in ['yes', 'true', '1'] else 'KONVENSIONAL'

            # Parse maturity date - handle both "MaturityDate" and "Maturity Date"
            maturity_date = row.get('MaturityDate') if 'MaturityDate' in row else row.get('Maturity Date')
            if pd.isna(maturity_date):
                continue
            if isinstance(maturity_date, pd.Timestamp):
                maturity_date = maturity_date.date()

            # Parse amount - handle both "DepositAmount" and "Amount"
            principal_amount = row.get('DepositAmount', row.get('Amount', 0))
            if pd.isna(principal_amount):
                principal_amount = 0

            asset = Asset(
                snapshot_id=snapshot.id,
                fund_id=fund_id,
                sub_fund=sub_fund,
                management_type=management_type,
                asset_type='DEPOSITO',
                security_id=str(row.get('CounterpartID', '')),
                security_name=str(row.get('CounterpartName', '')),
                maturity_date=maturity_date,
                principal_amount=Decimal(str(principal_amount)),
                coupon_amount=Decimal('0.00'),
                frequency_months=0,
                next_coupon_date=None,
                securities_type_desc=str(row.get('PlacementType', ''))
            )
            db.add(asset)
            total_assets += Decimal(str(principal_amount))
    
    # Process Bond
    if not bond_df.empty:
        for _, row in bond_df.iterrows():
            # Determine OPEX/CAPEX - FundID can be "OPEX", "CAPEX", or contain "JPOIP"
            fund_id = str(row.get('FundID', ''))
            sub_fund = 'OPEX' if (fund_id.upper() == 'OPEX' or 'JPOIP' in fund_id.upper()) else 'CAPEX'
            
            # Determine Syariah/Konvensional
            syariah_val = row.get('Syariah', False)
            if isinstance(syariah_val, bool):
                management_type = 'SYARIAH' if syariah_val else 'KONVENSIONAL'
            else:
                syariah_str = str(syariah_val).lower()
                management_type = 'SYARIAH' if syariah_str in ['yes', 'true', '1'] else 'KONVENSIONAL'

            # Parse dates - helper function
            def parse_date(val):
                if pd.isna(val):
                    return None
                if isinstance(val, pd.Timestamp):
                    return val.date()
                if isinstance(val, datetime):
                    return val.date()
                if isinstance(val, date):
                    return val
                return None

            # Handle both "MaturityDate" and "Maturity Date" column names
            maturity_date_col = 'MaturityDate' if 'MaturityDate' in row else 'Maturity Date'
            maturity_date = parse_date(row.get(maturity_date_col))
            if not maturity_date:
                continue

            # Handle both "NextCouponDate" and "Next Coupon Date" column names
            next_coupon_date_col = 'NextCouponDate' if 'NextCouponDate' in row else 'Next Coupon Date'
            next_coupon_date = parse_date(row.get(next_coupon_date_col))
            if not next_coupon_date:
                continue

            # Parse amounts - handle column name variations
            principal_amount = row.get('Balance', row.get('BalanceAmount', 0))
            if pd.isna(principal_amount):
                principal_amount = 0

            coupon_amount = row.get('NetCouponAmount', row.get('Net_Coupon_Amount', 0))
            if pd.isna(coupon_amount):
                coupon_amount = 0

            # Parse frequency - extract number from strings like "6 MONTHLY"
            frequency_months = row.get('Frequency', 0)
            if pd.isna(frequency_months):
                frequency_months = 0
            if isinstance(frequency_months, str):
                match = re.search(r'\d+', str(frequency_months))
                frequency_months = int(match.group()) if match else 0
            elif isinstance(frequency_months, (int, float)):
                frequency_months = int(frequency_months) if frequency_months else 0
            else:
                frequency_months = 0
            
            # Create Bond Maturity record (principal at maturity date)
            asset_maturity = Asset(
                snapshot_id=snapshot.id,
                fund_id=fund_id,
                sub_fund=sub_fund,
                management_type=management_type,
                asset_type='BOND',
                security_id=str(row.get('SecuritiesID', '')),
                security_name=str(row.get('SecuritiesName', '')),
                maturity_date=maturity_date,
                principal_amount=Decimal(str(principal_amount)),
                coupon_amount=Decimal('0.00'),
                frequency_months=frequency_months,
                next_coupon_date=next_coupon_date,
                securities_type_desc=str(row.get('SecuritiesTypeDescription', ''))
            )
            db.add(asset_maturity)
            total_assets += Decimal(str(principal_amount))
            
            # Generate recurring coupon payments from next_coupon_date until maturity
            if frequency_months > 0 and coupon_amount > 0:
                from dateutil.relativedelta import relativedelta
                # Ensure both dates are date objects (not datetime)
                if hasattr(next_coupon_date, 'date'):
                    current_coupon_date = next_coupon_date.date()
                else:
                    current_coupon_date = next_coupon_date
                    
                if hasattr(maturity_date, 'date'):
                    end_date = maturity_date.date()
                else:
                    end_date = maturity_date
                
                while current_coupon_date <= end_date:
                    asset_coupon = Asset(
                        snapshot_id=snapshot.id,
                        fund_id=fund_id,
                        sub_fund=sub_fund,
                        management_type=management_type,
                        asset_type='BOND_COUPON',
                        security_id=str(row.get('SecuritiesID', '')),
                        security_name=str(row.get('SecuritiesName', '')),
                        maturity_date=current_coupon_date,
                        principal_amount=Decimal('0.00'),
                        coupon_amount=Decimal(str(coupon_amount)),
                        frequency_months=frequency_months,
                        next_coupon_date=next_coupon_date,
                        securities_type_desc=str(row.get('SecuritiesTypeDescription', ''))
                    )
                    db.add(asset_coupon)
                    total_assets += Decimal(str(coupon_amount))
                    
                    # Move to next coupon date
                    current_coupon_date = current_coupon_date + relativedelta(months=frequency_months)
    
    # Update snapshot total
    snapshot.total_assets = total_assets
    db.commit()
    db.refresh(snapshot)
    
    return snapshot


@router.post("/upload", response_model=PortfolioSnapshotResponse)
async def upload_portfolio(
    file: UploadFile = File(...),
    clear_previous: bool = Query(False, description="Clear previous portfolio data before upload"),
    db: Session = Depends(get_db)
):
    """Upload portfolio data from Excel file."""
    try:
        # Clear previous data if requested
        if clear_previous:
            db.query(Asset).delete()
            db.query(PortfolioSnapshot).delete()
            db.commit()
        
        snapshot = parse_portfolio_excel(file, db)
        return snapshot
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/profile", response_model=MaturityProfileSummary)
async def get_maturity_profile(
    sub_fund: Optional[str] = Query(None, pattern="^(OPEX|CAPEX)$"),
    management_type: Optional[str] = Query(None, pattern="^(SYARIAH|KONVENSIONAL)$"),
    asset_type: Optional[str] = Query(None, pattern="^(DEPOSITO|BOND|BOND_COUPON)$"),
    group_by: str = Query("month", pattern="^(month|year)$"),
    db: Session = Depends(get_db)
):
    """
    Get maturity profile grouped by month or year.
    Shows principal and coupon amounts maturing in each period.
    """
    # Build filters
    filters = []

    if sub_fund:
        filters.append(Asset.sub_fund == sub_fund)

    if management_type:
        filters.append(Asset.management_type == management_type)

    if asset_type:
        filters.append(Asset.asset_type == asset_type)
    
    # Get the latest snapshot
    latest_snapshot = db.query(PortfolioSnapshot).order_by(PortfolioSnapshot.upload_date.desc()).first()
    if not latest_snapshot:
        return MaturityProfileSummary(data=[])
    
    filters.append(Asset.snapshot_id == latest_snapshot.id)
    
    # Query assets with maturity data
    query = db.query(
        Asset.maturity_date,
        Asset.asset_type,
        func.sum(Asset.principal_amount).label('total_principal'),
        func.sum(Asset.coupon_amount).label('total_coupon'),
        func.count(Asset.id).label('asset_count')
    ).filter(*filters).group_by(
        Asset.maturity_date,
        Asset.asset_type
    ).order_by(Asset.maturity_date)
    
    results = query.all()
    
    # Group by period (month or year)
    period_data = {}
    for row in results:
        if group_by == 'year':
            period = str(row.maturity_date.year)
        else:  # month
            period = f"{row.maturity_date.year}-{row.maturity_date.month:02d}"
        
        if period not in period_data:
            period_data[period] = {
                'maturity_amount': Decimal('0.00'),
                'coupon_amount': Decimal('0.00'),
                'asset_count': 0
            }
        
        period_data[period]['maturity_amount'] += row.total_principal or Decimal('0.00')
        period_data[period]['coupon_amount'] += row.total_coupon or Decimal('0.00')
        period_data[period]['asset_count'] += row.asset_count or 0
    
    # Build response
    profile_items = []
    total_principal = Decimal('0.00')
    
    for period in sorted(period_data.keys()):
        data = period_data[period]
        total_amount = data['maturity_amount'] + data['coupon_amount']
        total_principal += data['maturity_amount']
        
        profile_items.append(MaturityProfileItem(
            period=period,
            maturity_amount=data['maturity_amount'],
            coupon_amount=data['coupon_amount'],
            total_amount=total_amount,
            asset_count=data['asset_count']
        ))
    
    return MaturityProfileSummary(
        sub_fund=sub_fund,
        management_type=management_type,
        asset_type=asset_type,
        total_principal=total_principal,
        data=profile_items
    )


@router.get("/calendar", response_model=MaturityCalendarSummary)
async def get_maturity_calendar(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    sub_fund: Optional[str] = Query(None, pattern="^(OPEX|CAPEX)$"),
    management_type: Optional[str] = Query(None, pattern="^(SYARIAH|KONVENSIONAL)$"),
    asset_type: Optional[str] = Query(None, pattern="^(DEPOSITO|BOND|BOND_COUPON)$"),
    db: Session = Depends(get_db)
):
    """
    Get detailed maturity calendar with individual asset maturities.
    """
    # Build filters
    filters = []

    # Get the latest snapshot
    latest_snapshot = db.query(PortfolioSnapshot).order_by(PortfolioSnapshot.upload_date.desc()).first()
    if not latest_snapshot:
        return MaturityCalendarSummary(data=[])

    filters.append(Asset.snapshot_id == latest_snapshot.id)

    if start_date:
        filters.append(Asset.maturity_date >= start_date)

    if end_date:
        filters.append(Asset.maturity_date <= end_date)
    
    if sub_fund:
        filters.append(Asset.sub_fund == sub_fund)
    
    if management_type:
        filters.append(Asset.management_type == management_type)
    
    if asset_type:
        filters.append(Asset.asset_type == asset_type)
    
    # Query assets
    query = db.query(
        Asset.maturity_date,
        Asset.sub_fund,
        Asset.management_type,
        Asset.asset_type,
        Asset.security_id,
        Asset.security_name,
        Asset.principal_amount,
        Asset.coupon_amount
    ).filter(*filters).order_by(Asset.maturity_date, Asset.security_name)
    
    results = query.all()
    
    # Build response
    calendar_items = []
    total_amount = Decimal('0.00')
    
    for row in results:
        item_total = (row.principal_amount or Decimal('0.00')) + (row.coupon_amount or Decimal('0.00'))
        total_amount += item_total
        
        calendar_items.append(MaturityCalendarItem(
            maturity_date=row.maturity_date,
            sub_fund=row.sub_fund,
            management_type=row.management_type,
            asset_type=row.asset_type,
            security_id=row.security_id or '',
            security_name=row.security_name or '',
            principal_amount=row.principal_amount or Decimal('0.00'),
            coupon_amount=row.coupon_amount or Decimal('0.00'),
            total_amount=item_total
        ))
    
    return MaturityCalendarSummary(
        sub_fund=sub_fund,
        management_type=management_type,
        total_amount=total_amount,
        data=calendar_items
    )


@router.get("/summary")
async def get_portfolio_summary(
    db: Session = Depends(get_db)
):
    """Get summary of current portfolio."""
    # Get the latest snapshot
    latest_snapshot = db.query(PortfolioSnapshot).order_by(PortfolioSnapshot.upload_date.desc()).first()
    if not latest_snapshot:
        return {
            "total_assets": 0,
            "asset_count": 0,
            "last_updated": None
        }
    
    # Get asset counts by type
    deposito_count = db.query(func.count(Asset.id)).filter(
        Asset.snapshot_id == latest_snapshot.id,
        Asset.asset_type == 'DEPOSITO'
    ).scalar()
    
    bond_count = db.query(func.count(Asset.id)).filter(
        Asset.snapshot_id == latest_snapshot.id,
        Asset.asset_type == 'BOND'
    ).scalar()
    
    return {
        "total_assets": float(latest_snapshot.total_assets),
        "asset_count": deposito_count + bond_count,
        "deposito_count": deposito_count,
        "bond_count": bond_count,
        "last_updated": latest_snapshot.upload_date.isoformat()
    }


@router.get("/calendar/export")
async def export_maturity_calendar_to_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    sub_fund: Optional[str] = Query(None, pattern="^(OPEX|CAPEX)$"),
    management_type: Optional[str] = Query(None, pattern="^(SYARIAH|KONVENSIONAL)$"),
    asset_type: Optional[str] = Query(None, pattern="^(DEPOSITO|BOND|BOND_COUPON)$"),
    db: Session = Depends(get_db)
):
    """
    Export maturity calendar to Excel file.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Build filters
    filters = []

    # Get the latest snapshot
    latest_snapshot = db.query(PortfolioSnapshot).order_by(PortfolioSnapshot.upload_date.desc()).first()
    if not latest_snapshot:
        raise HTTPException(status_code=404, detail="No portfolio data found")

    filters.append(Asset.snapshot_id == latest_snapshot.id)

    if start_date:
        filters.append(Asset.maturity_date >= start_date)

    if end_date:
        filters.append(Asset.maturity_date <= end_date)

    if sub_fund:
        filters.append(Asset.sub_fund == sub_fund)

    if management_type:
        filters.append(Asset.management_type == management_type)

    if asset_type:
        filters.append(Asset.asset_type == asset_type)

    # Query assets
    query = db.query(
        Asset.maturity_date,
        Asset.sub_fund,
        Asset.management_type,
        Asset.asset_type,
        Asset.security_id,
        Asset.security_name,
        Asset.principal_amount,
        Asset.coupon_amount
    ).filter(*filters).order_by(Asset.maturity_date, Asset.security_name)

    results = query.all()

    # Build report data
    report_rows = []
    for row in results:
        report_rows.append({
            'Maturity Date': row.maturity_date,
            'Sub Fund': row.sub_fund,
            'Management Type': 'Syariah' if row.management_type == 'SYARIAH' else 'Conventional',
            'Asset Type': row.asset_type,
            'Security ID': row.security_id or '',
            'Security Name': row.security_name or '',
            'Principal Amount': float(row.principal_amount) if row.principal_amount else 0.0,
            'Coupon Amount': float(row.coupon_amount) if row.coupon_amount else 0.0,
            'Total Amount': float((row.principal_amount or 0) + (row.coupon_amount or 0))
        })

    # Create DataFrame
    df = pd.DataFrame(report_rows)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Main report sheet
        df.to_excel(writer, sheet_name='Maturity Calendar', index=False)

        # Get worksheet and format
        worksheet = writer.sheets['Maturity Calendar']

        # Format header
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format columns
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 20
        worksheet.column_dimensions['F'].width = 40
        worksheet.column_dimensions['G'].width = 18
        worksheet.column_dimensions['H'].width = 18
        worksheet.column_dimensions['I'].width = 18

        # Format number columns
        number_format = '#,##0.00'
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=7, max_col=9):
            for cell in row:
                cell.number_format = number_format

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=9):
            for cell in row:
                cell.border = thin_border

        # Add totals row
        if len(df) > 0:
            total_row = worksheet.max_row + 1
            worksheet.cell(row=total_row, column=1, value='TOTAL')
            for col in range(2, 7):
                worksheet.cell(row=total_row, column=col, value='')
            worksheet.cell(row=total_row, column=7, value=f'=SUM(G2:G{total_row-1})')
            worksheet.cell(row=total_row, column=8, value=f'=SUM(H2:H{total_row-1})')
            worksheet.cell(row=total_row, column=9, value=f'=SUM(I2:I{total_row-1})')

            # Format total row
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
            for cell in worksheet[total_row]:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

    output.seek(0)

    # Generate filename
    mgmt_label = management_type if management_type else 'All'
    subfund_label = sub_fund if sub_fund else 'All'
    asset_label = asset_type if asset_type else 'All'
    
    if start_date and end_date:
        date_label = f"{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}"
    else:
        date_label = 'All'
    
    filename = f"Maturity_Calendar_{date_label}_{subfund_label}_{mgmt_label}_{asset_label}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/profile/export")
async def export_maturity_profile_to_excel(
    sub_fund: Optional[str] = Query(None, pattern="^(OPEX|CAPEX)$"),
    management_type: Optional[str] = Query(None, pattern="^(SYARIAH|KONVENSIONAL)$"),
    asset_type: Optional[str] = Query(None, pattern="^(DEPOSITO|BOND|BOND_COUPON)$"),
    group_by: str = Query("month", pattern="^(month|year)$"),
    db: Session = Depends(get_db)
):
    """
    Export maturity profile to Excel file with summary and chart.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    # Build filters
    filters = []

    if sub_fund:
        filters.append(Asset.sub_fund == sub_fund)

    if management_type:
        filters.append(Asset.management_type == management_type)

    if asset_type:
        filters.append(Asset.asset_type == asset_type)

    # Get the latest snapshot
    latest_snapshot = db.query(PortfolioSnapshot).order_by(PortfolioSnapshot.upload_date.desc()).first()
    if not latest_snapshot:
        raise HTTPException(status_code=404, detail="No portfolio data found")

    filters.append(Asset.snapshot_id == latest_snapshot.id)

    # Query assets with maturity data
    query = db.query(
        Asset.maturity_date,
        Asset.asset_type,
        func.sum(Asset.principal_amount).label('total_principal'),
        func.sum(Asset.coupon_amount).label('total_coupon'),
        func.count(Asset.id).label('asset_count')
    ).filter(*filters).group_by(
        Asset.maturity_date,
        Asset.asset_type
    ).order_by(Asset.maturity_date)

    results = query.all()

    # Group by period (month or year)
    period_data = {}
    for row in results:
        if group_by == 'year':
            period = str(row.maturity_date.year)
        else:  # month
            period = f"{row.maturity_date.year}-{row.maturity_date.month:02d}"

        if period not in period_data:
            period_data[period] = {
                'maturity_amount': Decimal('0.00'),
                'coupon_amount': Decimal('0.00'),
                'asset_count': 0
            }

        period_data[period]['maturity_amount'] += row.total_principal or Decimal('0.00')
        period_data[period]['coupon_amount'] += row.total_coupon or Decimal('0.00')
        period_data[period]['asset_count'] += row.asset_count or 0

    # Build report data
    report_rows = []
    total_principal = Decimal('0.00')
    total_coupon = Decimal('0.00')

    for period in sorted(period_data.keys()):
        data = period_data[period]
        total_amount = data['maturity_amount'] + data['coupon_amount']
        total_principal += data['maturity_amount']
        total_coupon += data['coupon_amount']

        report_rows.append({
            'Period': period,
            'Maturity Amount': float(data['maturity_amount']),
            'Coupon Amount': float(data['coupon_amount']),
            'Total Amount': float(total_amount),
            'Asset Count': data['asset_count']
        })

    # Create DataFrame
    df = pd.DataFrame(report_rows)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Summary sheet
        df.to_excel(writer, sheet_name='Maturity Profile', index=False)

        # Get worksheet and format
        worksheet = writer.sheets['Maturity Profile']

        # Format header
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format columns
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 15

        # Format number columns
        number_format = '#,##0.00'
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = number_format

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border

        # Add totals row
        if len(df) > 0:
            total_row = worksheet.max_row + 1
            worksheet.cell(row=total_row, column=1, value='TOTAL')
            worksheet.cell(row=total_row, column=2, value=f'=SUM(B2:B{total_row-1})')
            worksheet.cell(row=total_row, column=3, value=f'=SUM(C2:C{total_row-1})')
            worksheet.cell(row=total_row, column=4, value=f'=SUM(D2:D{total_row-1})')
            worksheet.cell(row=total_row, column=5, value=f'=SUM(E2:E{total_row-1})')

            # Format total row
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
            for cell in worksheet[total_row]:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Maturity Profile by Period"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Period"

        # Add data for chart (Maturity Amount and Coupon Amount)
        data_ref = Reference(worksheet, min_col=2, min_row=1, max_col=3, max_row=len(df) + 1)
        cats_ref = Reference(worksheet, min_col=1, min_row=2, max_row=len(df) + 1)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        
        # Position chart
        worksheet.add_chart(chart, "G2")

    output.seek(0)

    # Generate filename
    mgmt_label = management_type if management_type else 'All'
    subfund_label = sub_fund if sub_fund else 'All'
    asset_label = asset_type if asset_type else 'All'
    group_label = 'Yearly' if group_by == 'year' else 'Monthly'
    
    filename = f"Maturity_Profile_{group_label}_{subfund_label}_{mgmt_label}_{asset_label}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
