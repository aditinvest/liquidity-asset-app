from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from typing import Optional, Dict, Any, List, Tuple
from datetime import date
from decimal import Decimal
import pandas as pd
from io import BytesIO

from ..database import get_db
from ..models.models import CFProjection, ManualInput, CFRealization

router = APIRouter()


def get_projection_data(db: Session, month: int, year: int, sub_fund: Optional[str], management_type: Optional[str]):
    """Helper function to get projection data with rolling balances."""
    
    # Build query filters for projections
    projection_filters = [
        extract('month', CFProjection.transaction_date) == month,
        extract('year', CFProjection.transaction_date) == year
    ]

    if sub_fund:
        projection_filters.append(CFProjection.sub_fund == sub_fund)

    if management_type:
        projection_filters.append(CFProjection.management_type == management_type)

    # Get projection data grouped by date
    projection_data = db.query(
        CFProjection.transaction_date,
        CFProjection.sub_fund,
        CFProjection.management_type,
        func.sum(CFProjection.amount).filter(CFProjection.instrument_type == 'DEPOSITO').label('deposito_in'),
        func.sum(CFProjection.amount).filter(CFProjection.instrument_type == 'BOND_COUPON').label('bond_coupon_in'),
        func.sum(CFProjection.amount).filter(CFProjection.instrument_type == 'BOND_MATURITY').label('bond_maturity_in')
    ).filter(*projection_filters).group_by(
        CFProjection.transaction_date,
        CFProjection.sub_fund,
        CFProjection.management_type
    ).all()

    # Get CF Out (manual inputs) for the same period
    cf_out_filters = [
        extract('month', ManualInput.transaction_date) == month,
        extract('year', ManualInput.transaction_date) == year,
        ManualInput.input_type == 'CF_OUT'
    ]

    if sub_fund:
        cf_out_filters.append(ManualInput.sub_fund == sub_fund)

    if management_type:
        cf_out_filters.append(ManualInput.management_type == management_type)

    cf_out_data = db.query(
        ManualInput.transaction_date,
        ManualInput.sub_fund,
        ManualInput.management_type,
        func.sum(ManualInput.amount).label('cf_out')
    ).filter(*cf_out_filters).group_by(
        ManualInput.transaction_date,
        ManualInput.sub_fund,
        ManualInput.management_type
    ).all()

    # Get Opening Balance (Saldo Awal)
    saldo_awal_filters = [
        ManualInput.input_type == 'SALDO_AWAL'
    ]

    if sub_fund:
        saldo_awal_filters.append(ManualInput.sub_fund == sub_fund)

    if management_type:
        saldo_awal_filters.append(ManualInput.management_type == management_type)

    month_start = date(year, month, 1)
    saldo_awal_filters.append(ManualInput.transaction_date <= month_start)

    saldo_awal_data = db.query(
        ManualInput.sub_fund,
        ManualInput.management_type,
        func.max(ManualInput.transaction_date).label('latest_date'),
        func.sum(ManualInput.amount).label('opening_balance')
    ).filter(*saldo_awal_filters).group_by(
        ManualInput.sub_fund,
        ManualInput.management_type
    ).all()

    # Create lookups
    cf_out_lookup = {}
    for row in cf_out_data:
        key = (row.transaction_date, row.sub_fund, row.management_type)
        cf_out_lookup[key] = row.cf_out or Decimal('0.00')

    opening_balance_lookup = {}
    for row in saldo_awal_data:
        key = (row.sub_fund, row.management_type)
        opening_balance_lookup[key] = row.opening_balance or Decimal('0.00')

    # Collect all unique keys
    all_keys: set = set()
    projection_lookup = {}

    for row in projection_data:
        key = (row.transaction_date, row.sub_fund, row.management_type)
        all_keys.add(key)
        projection_lookup[key] = {
            'deposito_in': row.deposito_in or Decimal('0.00'),
            'bond_coupon_in': row.bond_coupon_in or Decimal('0.00'),
            'bond_maturity_in': row.bond_maturity_in or Decimal('0.00')
        }

    for row in cf_out_data:
        key = (row.transaction_date, row.sub_fund, row.management_type)
        all_keys.add(key)

    # Handle "All" management type
    if management_type is None:
        grouped_by_date: Dict[Tuple[str, date], Dict[str, Any]] = {}
        
        for key in all_keys:
            transaction_date, sub_fund_val, mgmt_type = key
            group_key = (sub_fund_val, transaction_date)
            
            if group_key not in grouped_by_date:
                grouped_by_date[group_key] = {
                    'deposito_in': Decimal('0.00'),
                    'bond_coupon_in': Decimal('0.00'),
                    'bond_maturity_in': Decimal('0.00'),
                    'cf_out': Decimal('0.00')
                }
            
            proj = projection_lookup.get(key, {
                'deposito_in': Decimal('0.00'),
                'bond_coupon_in': Decimal('0.00'),
                'bond_maturity_in': Decimal('0.00')
            })
            
            grouped_by_date[group_key]['deposito_in'] += proj['deposito_in']
            grouped_by_date[group_key]['bond_coupon_in'] += proj['bond_coupon_in']
            grouped_by_date[group_key]['bond_maturity_in'] += proj['bond_maturity_in']
            grouped_by_date[group_key]['cf_out'] += cf_out_lookup.get(key, Decimal('0.00'))
        
        sorted_dates = sorted(grouped_by_date.keys(), key=lambda x: x[1])
        running_balance_by_subfund: Dict[str, Decimal] = {}
        report_rows = []
        
        for group_key in sorted_dates:
            sub_fund_val, transaction_date = group_key
            data = grouped_by_date[group_key]
            
            if sub_fund_val not in running_balance_by_subfund:
                opening_balance = Decimal('0.00')
                for mgmt_type in ['SYARIAH', 'KONVENSIONAL']:
                    opening_balance += opening_balance_lookup.get((sub_fund_val, mgmt_type), Decimal('0.00'))
            else:
                opening_balance = running_balance_by_subfund[sub_fund_val]
            
            deposito_in = data['deposito_in']
            bond_coupon_in = data['bond_coupon_in']
            bond_maturity_in = data['bond_maturity_in']
            cf_out = data['cf_out']
            
            total_in = deposito_in + bond_coupon_in + bond_maturity_in
            net_cashflow = total_in - cf_out
            ending_balance = opening_balance + net_cashflow
            
            running_balance_by_subfund[sub_fund_val] = ending_balance
            
            report_rows.append({
                'date': transaction_date,
                'sub_fund': sub_fund_val,
                'management_type': 'ALL',
                'deposito_in': float(deposito_in),
                'bond_coupon': float(bond_coupon_in),
                'bond_maturity': float(bond_maturity_in),
                'total_in': float(total_in),
                'cf_out': float(cf_out),
                'net_cf': float(net_cashflow),
                'opening_balance': float(opening_balance),
                'ending_balance': float(ending_balance)
            })
        
        return report_rows
    else:
        grouped_keys: Dict[Tuple[str, str], List[Tuple[date, str, str]]] = {}
        for key in all_keys:
            transaction_date, sub_fund_val, mgmt_type = key
            group_key = (sub_fund_val, mgmt_type)
            if group_key not in grouped_keys:
                grouped_keys[group_key] = []
            grouped_keys[group_key].append(key)

        for group_key in grouped_keys:
            grouped_keys[group_key].sort(key=lambda x: x[0])

        running_balance: Dict[Tuple[str, str], Decimal] = {}
        report_rows = []

        for group_key in sorted(grouped_keys.keys()):
            keys_list = grouped_keys[group_key]

            for key in keys_list:
                transaction_date, sub_fund_val, mgmt_type = key
                opening_key = (sub_fund_val, mgmt_type)

                proj = projection_lookup.get(key, {
                    'deposito_in': Decimal('0.00'),
                    'bond_coupon_in': Decimal('0.00'),
                    'bond_maturity_in': Decimal('0.00')
                })

                deposito_in = proj['deposito_in']
                bond_coupon_in = proj['bond_coupon_in']
                bond_maturity_in = proj['bond_maturity_in']
                cf_out = cf_out_lookup.get(key, Decimal('0.00'))

                if opening_key not in running_balance:
                    opening_balance = opening_balance_lookup.get(opening_key, Decimal('0.00'))
                else:
                    opening_balance = running_balance[opening_key]

                total_in = deposito_in + bond_coupon_in + bond_maturity_in
                net_cashflow = total_in - cf_out
                ending_balance = opening_balance + net_cashflow

                running_balance[opening_key] = ending_balance

                report_rows.append({
                    'date': transaction_date,
                    'sub_fund': sub_fund_val,
                    'management_type': mgmt_type,
                    'deposito_in': float(deposito_in),
                    'bond_coupon': float(bond_coupon_in),
                    'bond_maturity': float(bond_maturity_in),
                    'total_in': float(total_in),
                    'cf_out': float(cf_out),
                    'net_cf': float(net_cashflow),
                    'opening_balance': float(opening_balance),
                    'ending_balance': float(ending_balance)
                })
        
        return report_rows


@router.get("/pivot")
async def get_projection_pivot(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    sub_fund: Optional[str] = Query(None, pattern="^(OPEX|CAPEX)$"),
    management_type: Optional[str] = Query(None, pattern="^(SYARIAH|KONVENSIONAL)$"),
    db: Session = Depends(get_db)
):
    """Get projection data in pivot format."""
    return get_projection_data(db, month, year, sub_fund, management_type)


@router.get("/export")
async def export_projection_to_excel(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    sub_fund: Optional[str] = Query(None, pattern="^(OPEX|CAPEX)$"),
    management_type: Optional[str] = Query(None, pattern="^(SYARIAH|KONVENSIONAL)$"),
    db: Session = Depends(get_db)
):
    """
    Export cash flow projection report to Excel file.
    """
    from fastapi.responses import StreamingResponse

    # Build query filters for projections
    projection_filters = [
        extract('month', CFProjection.transaction_date) == month,
        extract('year', CFProjection.transaction_date) == year
    ]

    if sub_fund:
        projection_filters.append(CFProjection.sub_fund == sub_fund)

    if management_type:
        projection_filters.append(CFProjection.management_type == management_type)

    # Get projection data grouped by date
    projection_data = db.query(
        CFProjection.transaction_date,
        CFProjection.sub_fund,
        CFProjection.management_type,
        func.sum(CFProjection.amount).filter(CFProjection.instrument_type == 'DEPOSITO').label('deposito_in'),
        func.sum(CFProjection.amount).filter(CFProjection.instrument_type == 'BOND_COUPON').label('bond_coupon_in'),
        func.sum(CFProjection.amount).filter(CFProjection.instrument_type == 'BOND_MATURITY').label('bond_maturity_in')
    ).filter(*projection_filters).group_by(
        CFProjection.transaction_date,
        CFProjection.sub_fund,
        CFProjection.management_type
    ).all()

    # Get CF Out (manual inputs) for the same period
    cf_out_filters = [
        extract('month', ManualInput.transaction_date) == month,
        extract('year', ManualInput.transaction_date) == year,
        ManualInput.input_type == 'CF_OUT'
    ]

    if sub_fund:
        cf_out_filters.append(ManualInput.sub_fund == sub_fund)

    if management_type:
        cf_out_filters.append(ManualInput.management_type == management_type)

    cf_out_data = db.query(
        ManualInput.transaction_date,
        ManualInput.sub_fund,
        ManualInput.management_type,
        func.sum(ManualInput.amount).label('cf_out')
    ).filter(*cf_out_filters).group_by(
        ManualInput.transaction_date,
        ManualInput.sub_fund,
        ManualInput.management_type
    ).all()

    # Get Opening Balance (Saldo Awal)
    saldo_awal_filters = [
        ManualInput.input_type == 'SALDO_AWAL'
    ]

    if sub_fund:
        saldo_awal_filters.append(ManualInput.sub_fund == sub_fund)

    if management_type:
        saldo_awal_filters.append(ManualInput.management_type == management_type)

    month_start = date(year, month, 1)
    saldo_awal_filters.append(ManualInput.transaction_date <= month_start)

    saldo_awal_data = db.query(
        ManualInput.sub_fund,
        ManualInput.management_type,
        func.max(ManualInput.transaction_date).label('latest_date'),
        func.sum(ManualInput.amount).label('opening_balance')
    ).filter(*saldo_awal_filters).group_by(
        ManualInput.sub_fund,
        ManualInput.management_type
    ).all()

    # Create lookups
    cf_out_lookup = {}
    for row in cf_out_data:
        key = (row.transaction_date, row.sub_fund, row.management_type)
        cf_out_lookup[key] = row.cf_out or Decimal('0.00')

    opening_balance_lookup = {}
    for row in saldo_awal_data:
        key = (row.sub_fund, row.management_type)
        opening_balance_lookup[key] = row.opening_balance or Decimal('0.00')

    # Collect all unique keys
    all_keys: set = set()
    projection_lookup = {}

    for row in projection_data:
        key = (row.transaction_date, row.sub_fund, row.management_type)
        all_keys.add(key)
        projection_lookup[key] = {
            'deposito_in': row.deposito_in or Decimal('0.00'),
            'bond_coupon_in': row.bond_coupon_in or Decimal('0.00'),
            'bond_maturity_in': row.bond_maturity_in or Decimal('0.00')
        }

    for row in cf_out_data:
        key = (row.transaction_date, row.sub_fund, row.management_type)
        all_keys.add(key)

    # Handle "All" management type
    if management_type is None:
        grouped_by_date: Dict[Tuple[str, date], Dict[str, Any]] = {}
        
        for key in all_keys:
            transaction_date, sub_fund_val, mgmt_type = key
            group_key = (sub_fund_val, transaction_date)
            
            if group_key not in grouped_by_date:
                grouped_by_date[group_key] = {
                    'deposito_in': Decimal('0.00'),
                    'bond_coupon_in': Decimal('0.00'),
                    'bond_maturity_in': Decimal('0.00'),
                    'cf_out': Decimal('0.00')
                }
            
            proj = projection_lookup.get(key, {
                'deposito_in': Decimal('0.00'),
                'bond_coupon_in': Decimal('0.00'),
                'bond_maturity_in': Decimal('0.00')
            })
            
            grouped_by_date[group_key]['deposito_in'] += proj['deposito_in']
            grouped_by_date[group_key]['bond_coupon_in'] += proj['bond_coupon_in']
            grouped_by_date[group_key]['bond_maturity_in'] += proj['bond_maturity_in']
            grouped_by_date[group_key]['cf_out'] += cf_out_lookup.get(key, Decimal('0.00'))
        
        sorted_dates = sorted(grouped_by_date.keys(), key=lambda x: x[1])
        running_balance_by_subfund: Dict[str, Decimal] = {}
        report_rows = []
        
        for group_key in sorted_dates:
            sub_fund_val, transaction_date = group_key
            data = grouped_by_date[group_key]
            
            if sub_fund_val not in running_balance_by_subfund:
                opening_balance = Decimal('0.00')
                for mgmt_type in ['SYARIAH', 'KONVENSIONAL']:
                    opening_balance += opening_balance_lookup.get((sub_fund_val, mgmt_type), Decimal('0.00'))
            else:
                opening_balance = running_balance_by_subfund[sub_fund_val]
            
            deposito_in = data['deposito_in']
            bond_coupon_in = data['bond_coupon_in']
            bond_maturity_in = data['bond_maturity_in']
            cf_out = data['cf_out']
            
            total_in = deposito_in + bond_coupon_in + bond_maturity_in
            net_cashflow = total_in - cf_out
            ending_balance = opening_balance + net_cashflow
            
            running_balance_by_subfund[sub_fund_val] = ending_balance
            
            report_rows.append({
                'Date': transaction_date,
                'Sub Fund': sub_fund_val,
                'Management Type': 'All',
                'Opening Balance': float(opening_balance),
                'Deposito In': float(deposito_in),
                'Bond Coupon': float(bond_coupon_in),
                'Bond Maturity': float(bond_maturity_in),
                'Total In': float(total_in),
                'CF Out': float(cf_out),
                'Net CF': float(net_cashflow),
                'Ending Balance': float(ending_balance)
            })
    else:
        grouped_keys: Dict[Tuple[str, str], List[Tuple[date, str, str]]] = {}
        for key in all_keys:
            transaction_date, sub_fund_val, mgmt_type = key
            group_key = (sub_fund_val, mgmt_type)
            if group_key not in grouped_keys:
                grouped_keys[group_key] = []
            grouped_keys[group_key].append(key)

        for group_key in grouped_keys:
            grouped_keys[group_key].sort(key=lambda x: x[0])

        running_balance: Dict[Tuple[str, str], Decimal] = {}
        report_rows = []

        for group_key in sorted(grouped_keys.keys()):
            keys_list = grouped_keys[group_key]

            for key in keys_list:
                transaction_date, sub_fund_val, mgmt_type = key
                opening_key = (sub_fund_val, mgmt_type)

                proj = projection_lookup.get(key, {
                    'deposito_in': Decimal('0.00'),
                    'bond_coupon_in': Decimal('0.00'),
                    'bond_maturity_in': Decimal('0.00')
                })

                deposito_in = proj['deposito_in']
                bond_coupon_in = proj['bond_coupon_in']
                bond_maturity_in = proj['bond_maturity_in']
                cf_out = cf_out_lookup.get(key, Decimal('0.00'))

                if opening_key not in running_balance:
                    opening_balance = opening_balance_lookup.get(opening_key, Decimal('0.00'))
                else:
                    opening_balance = running_balance[opening_key]

                total_in = deposito_in + bond_coupon_in + bond_maturity_in
                net_cashflow = total_in - cf_out
                ending_balance = opening_balance + net_cashflow

                running_balance[opening_key] = ending_balance

                report_rows.append({
                    'Date': transaction_date,
                    'Sub Fund': sub_fund_val,
                    'Management Type': 'Syariah' if mgmt_type == 'SYARIAH' else 'Conventional',
                    'Opening Balance': float(opening_balance),
                    'Deposito In': float(deposito_in),
                    'Bond Coupon': float(bond_coupon_in),
                    'Bond Maturity': float(bond_maturity_in),
                    'Total In': float(total_in),
                    'CF Out': float(cf_out),
                    'Net CF': float(net_cashflow),
                    'Ending Balance': float(ending_balance)
                })

    # Sort by date
    report_rows.sort(key=lambda x: x['Date'])

    # Create DataFrame
    df = pd.DataFrame(report_rows)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Main report sheet
        df.to_excel(writer, sheet_name='Cash Flow Projection', index=False)
        
        # Get worksheet and format
        worksheet = writer.sheets['Cash Flow Projection']
        
        # Format header
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Format columns
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 18
        worksheet.column_dimensions['E'].width = 18
        worksheet.column_dimensions['F'].width = 18
        worksheet.column_dimensions['G'].width = 18
        worksheet.column_dimensions['H'].width = 18
        worksheet.column_dimensions['I'].width = 18
        worksheet.column_dimensions['J'].width = 18
        worksheet.column_dimensions['K'].width = 18
        
        # Format number columns
        number_format = '#,##0.00'
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=11):
            for cell in row:
                cell.number_format = number_format
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=11):
            for cell in row:
                cell.border = thin_border
        
        # Add totals row
        if len(df) > 0:
            total_row = worksheet.max_row + 1
            worksheet.cell(row=total_row, column=1, value='TOTAL')
            worksheet.cell(row=total_row, column=2, value='')
            worksheet.cell(row=total_row, column=3, value='')
            worksheet.cell(row=total_row, column=4, value='')  # Opening Balance - empty
            worksheet.cell(row=total_row, column=5, value=f'=SUM(E2:E{total_row-1})')  # Deposito In
            worksheet.cell(row=total_row, column=6, value=f'=SUM(F2:F{total_row-1})')  # Bond Coupon
            worksheet.cell(row=total_row, column=7, value=f'=SUM(G2:G{total_row-1})')  # Bond Maturity
            worksheet.cell(row=total_row, column=8, value=f'=SUM(H2:H{total_row-1})')  # Total In
            worksheet.cell(row=total_row, column=9, value=f'=SUM(I2:I{total_row-1})')  # CF Out
            worksheet.cell(row=total_row, column=10, value=f'=SUM(J2:J{total_row-1})')  # Net CF
            worksheet.cell(row=total_row, column=11, value='')  # Ending Balance - empty
            
            # Format total row
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
            for cell in worksheet[total_row]:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

    output.seek(0)
    
    # Generate filename
    mgmt_label = management_type if management_type else 'All'
    subfund_label = sub_fund if sub_fund else 'All'
    filename = f"CF_Projection_{year}_{month:02d}_{subfund_label}_{mgmt_label}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/realization/export")
async def export_realization_to_excel(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2020),
    sub_fund: Optional[str] = Query(None, pattern="^(OPEX|CAPEX)$"),
    management_type: Optional[str] = Query(None, pattern="^(SYARIAH|KONVENSIONAL)$"),
    db: Session = Depends(get_db)
):
    """
    Export cash flow realization report to Excel file.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Build filters
    filters = []

    if month:
        filters.append(extract('month', CFRealization.transaction_date) == month)

    if year:
        filters.append(extract('year', CFRealization.transaction_date) == year)

    if sub_fund:
        filters.append(CFRealization.sub_fund == sub_fund)

    if management_type:
        filters.append(CFRealization.management_type == management_type)

    # Get realization data
    query = db.query(
        CFRealization.transaction_date,
        CFRealization.sub_fund,
        CFRealization.management_type,
        CFRealization.transaction_category,
        CFRealization.instrument_name,
        CFRealization.proceed_amount
    )

    if filters:
        query = query.filter(*filters)

    results = query.order_by(CFRealization.transaction_date).all()

    # Build report data
    report_rows = []
    for row in results:
        report_rows.append({
            'Date': row.transaction_date,
            'Sub Fund': row.sub_fund,
            'Management Type': 'Syariah' if row.management_type == 'SYARIAH' else 'Conventional',
            'Transaction Category': row.transaction_category,
            'Instrument Name': row.instrument_name or '',
            'Proceed Amount': float(row.proceed_amount or Decimal('0.00'))
        })

    # Sort by date
    report_rows.sort(key=lambda x: x['Date'])

    # Create DataFrame
    df = pd.DataFrame(report_rows)

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Main report sheet
        df.to_excel(writer, sheet_name='Cash Flow Realization', index=False)

        # Get worksheet and format
        worksheet = writer.sheets['Cash Flow Realization']

        # Format header
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format columns
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 45
        worksheet.column_dimensions['F'].width = 20

        # Format amount column
        number_format = '#,##0.00'
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.number_format = number_format

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border

        # Add summary section
        if len(df) > 0:
            # Add totals by category
            summary_start = worksheet.max_row + 2
            worksheet.cell(row=summary_start, column=1, value='SUMMARY BY CATEGORY')
            summary_font = Font(bold=True, size=12)
            worksheet.cell(row=summary_start, column=1).font = summary_font

            # Calculate totals by category
            category_totals = df.groupby('Transaction Category')['Proceed Amount'].sum().reset_index()
            category_totals = category_totals.sort_values('Proceed Amount', ascending=False)

            summary_row = summary_start + 1
            worksheet.cell(row=summary_row, column=1, value='Category')
            worksheet.cell(row=summary_row, column=2, value='Total Amount')

            # Format header row
            for cell in worksheet[summary_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
                cell.border = thin_border

            # Add category data
            for idx, cat_row in category_totals.iterrows():
                summary_row += 1
                worksheet.cell(row=summary_row, column=1, value=cat_row['Transaction Category'])
                worksheet.cell(row=summary_row, column=2, value=cat_row['Proceed Amount'])
                for cell in worksheet[summary_row]:
                    cell.border = thin_border
                    if cell.column == 2:
                        cell.number_format = number_format

            # Add grand total
            summary_row += 1
            worksheet.cell(row=summary_row, column=1, value='GRAND TOTAL')
            worksheet.cell(row=summary_row, column=2, value=f'=SUM(B{summary_start + 2}:B{summary_row - 1})')
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            for cell in worksheet[summary_row]:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 20

    output.seek(0)

    # Generate filename
    mgmt_label = management_type if management_type else 'All'
    subfund_label = sub_fund if sub_fund else 'All'
    month_str = f'{year}_{month:02d}' if month and year else 'All'
    filename = f"CF_Realization_{month_str}_{subfund_label}_{mgmt_label}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/realization/export-pivot")
async def export_realization_pivot_to_excel(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2020),
    sub_fund: Optional[str] = Query(None, pattern="^(OPEX|CAPEX)$"),
    management_type: Optional[str] = Query(None, pattern="^(SYARIAH|KONVENSIONAL)$"),
    db: Session = Depends(get_db)
):
    """
    Export cash flow realization report in pivot format to Excel file.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Build filters
    filters = []

    if month:
        filters.append(extract('month', CFRealization.transaction_date) == month)

    if year:
        filters.append(extract('year', CFRealization.transaction_date) == year)

    if sub_fund:
        filters.append(CFRealization.sub_fund == sub_fund)

    if management_type:
        filters.append(CFRealization.management_type == management_type)

    # Get all unique dates
    dates_query = db.query(CFRealization.transaction_date).filter(*filters).distinct()
    dates = [row.transaction_date for row in dates_query.order_by(CFRealization.transaction_date)]

    # Get all unique category-instrument combinations
    combinations_query = db.query(
        CFRealization.transaction_category,
        CFRealization.instrument_name
    ).filter(*filters).distinct()

    combinations = [
        (row.transaction_category, row.instrument_name)
        for row in combinations_query.order_by(
            CFRealization.transaction_category,
            CFRealization.instrument_name
        )
    ]

    # Get pivot data - SUM by date, category, and instrument
    pivot_data = {}
    for d in dates:
        pivot_data[d] = {}
        for combo in combinations:
            pivot_data[d][combo] = Decimal('0.00')

    # Fill in actual values
    data_query = db.query(
        CFRealization.transaction_date,
        CFRealization.transaction_category,
        CFRealization.instrument_name,
        func.sum(CFRealization.proceed_amount).label('total_amount')
    ).filter(*filters).group_by(
        CFRealization.transaction_date,
        CFRealization.transaction_category,
        CFRealization.instrument_name
    )

    for row in data_query:
        key = (row.transaction_category, row.instrument_name or '')
        if row.transaction_date in pivot_data and key in pivot_data[row.transaction_date]:
            pivot_data[row.transaction_date][key] = row.total_amount or Decimal('0.00')

    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create pivot table data
        pivot_rows = []
        for combo in combinations:
            row_data = {
                'Transaction Category': combo[0],
                'Instrument Name': combo[1] or ''
            }
            for d in dates:
                row_data[str(d)] = float(pivot_data[d][combo])
            pivot_rows.append(row_data)

        df = pd.DataFrame(pivot_rows)

        # Main report sheet
        df.to_excel(writer, sheet_name='Pivot Table', index=False)

        # Get worksheet and format
        worksheet = writer.sheets['Pivot Table']

        # Format header
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format columns
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 45

        # Format date columns
        number_format = '#,##0.00'
        for col_idx in range(3, worksheet.max_column + 1):
            worksheet.column_dimensions[chr(64 + col_idx)].width = 15
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = number_format

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border

        # Add totals row
        if len(df) > 0:
            total_row = worksheet.max_row + 1
            worksheet.cell(row=total_row, column=1, value='TOTAL')
            worksheet.cell(row=total_row, column=2, value='')
            for col_idx in range(3, worksheet.max_column + 1):
                col_letter = chr(64 + col_idx)
                worksheet.cell(row=total_row, column=col_idx, value=f'=SUM({col_letter}2:{col_letter}{total_row-1})')

            # Format total row
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
            for cell in worksheet[total_row]:
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.column >= 3:
                    cell.number_format = number_format

    output.seek(0)

    # Generate filename
    mgmt_label = management_type if management_type else 'All'
    subfund_label = sub_fund if sub_fund else 'All'
    month_str = f'{year}_{month:02d}' if month and year else 'All'
    filename = f"CF_Realization_Pivot_{month_str}_{subfund_label}_{mgmt_label}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
